"""
LinkedIn Job Scraper for Italy
===============================
Scrapes geo/GIS jobs in Italy using Apify's LinkedIn Jobs Scraper.
"""

import json
import os
import time
import requests
import openpyxl
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import urlencode
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
#  CONFIGURATION FOR ITALY
# ─────────────────────────────────────────────

TOKEN_ENV_VAR = "APIFY_API_TOKEN"
TOKEN_FILE = Path(__file__).with_name(".env")
TOKEN_PLACEHOLDER = "apify_api_XXXXXXXXXXXX"
GOOGLE_CLIENT_SECRET_FILE = Path(__file__).with_name("google_client_secret.json")
GOOGLE_TOKEN_FILE = Path(__file__).with_name("google_token.json")
GOOGLE_SPREADSHEET_ID_FILE = Path(__file__).with_name("google_spreadsheet_id.txt")
GOOGLE_SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
]


def load_local_env() -> dict[str, str]:
    """Load simple KEY=value settings from local .env."""
    values = {}
    if not TOKEN_FILE.exists():
        return values

    for line in TOKEN_FILE.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        if line.startswith("export "):
            line = line[len("export "):].strip()
        if "=" not in line:
            continue

        key, value = line.split("=", 1)
        values[key.strip()] = value.strip().strip("\"'")

    return values


LOCAL_ENV = load_local_env()


def load_setting(name: str, default: str = "") -> str:
    return os.environ.get(name, LOCAL_ENV.get(name, default)).strip()


def load_int_setting(name: str, default: int) -> int:
    value = load_setting(name, str(default))
    try:
        return int(value)
    except ValueError:
        print(f"⚠ Invalid integer for {name}='{value}', using {default}.")
        return default


def load_bool_setting(name: str, default: bool) -> bool:
    value = load_setting(name, str(default).lower()).lower()
    if value in {"1", "true", "yes", "y", "on"}:
        return True
    if value in {"0", "false", "no", "n", "off"}:
        return False
    print(f"⚠ Invalid boolean for {name}='{value}', using {default}.")
    return default


def load_apify_token() -> str:
    return load_setting(TOKEN_ENV_VAR)


APIFY_API_TOKEN = load_apify_token()

# ITALY-SPECIFIC SETTINGS
LOCATION = "Italy"
GEO_ID = "103644278"  # Geo ID for Italy
PUBLISHED_AT = "r86400"  # Last 24 hours
EXPERIENCE_LEVELS = ["1", "2"]  # Internship, Entry level
CONTRACT_TYPES = ["F", "P", "I"]  # Full-time, Part-time, Internship
SCRAPE_COMPANY_DETAILS = load_bool_setting("JOBSCRAPER_SCRAPE_COMPANY_DETAILS", False)
USE_INCOGNITO_MODE = load_bool_setting("JOBSCRAPER_USE_INCOGNITO_MODE", True)
SPLIT_BY_LOCATION = load_bool_setting("JOBSCRAPER_SPLIT_BY_LOCATION", False)
SPLIT_COUNTRY = "IT"
EXCLUDED_TITLE_TERMS = ["Stage", "Tirocinio"]  # Words to exclude (internships if unwanted)

SCRAPER_TIMEZONE = load_setting("JOBSCRAPER_TIMEZONE", "Europe/Rome")
POSTED_TIMEZONE = load_setting("JOBSCRAPER_POSTED_TIMEZONE", "Europe/Rome")

try:
    SCRAPER_TZ = ZoneInfo(SCRAPER_TIMEZONE)
    POSTED_TZ = ZoneInfo(POSTED_TIMEZONE)
except ZoneInfoNotFoundError as e:
    raise RuntimeError(f"Timezone not available: {e}")

RUN_STARTED_AT_UTC = datetime.now(timezone.utc)
RUN_STARTED_AT = RUN_STARTED_AT_UTC.astimezone(SCRAPER_TZ)
RUN_SHEET_NAME = RUN_STARTED_AT.strftime("%Y-%m-%d %H-%M-%S")

# Apify actor
LINKEDIN_ACTOR_ID = "curious_coder~linkedin-jobs-scraper"

# Max jobs per search
MAX_RESULTS_PER_SEARCH = load_int_setting("JOBSCRAPER_MAX_RESULTS_PER_SEARCH", 500)
SEARCH_CONCURRENCY = max(1, load_int_setting("JOBSCRAPER_SEARCH_CONCURRENCY", 10))

# Apify settings
APIFY_RUN_MEMORY_MB = max(128, load_int_setting("APIFY_RUN_MEMORY_MB", 512))
APIFY_RUN_TIMEOUT_SECONDS = max(60, load_int_setting("APIFY_RUN_TIMEOUT_SECONDS", 300))
APIFY_CLIENT_TIMEOUT_SECONDS = max(
    APIFY_RUN_TIMEOUT_SECONDS + 30,
    load_int_setting("APIFY_CLIENT_TIMEOUT_SECONDS", APIFY_RUN_TIMEOUT_SECONDS + 60),
)

DELAY_BETWEEN_REQUESTS = max(0, load_int_setting("JOBSCRAPER_DELAY_BETWEEN_REQUESTS", 0))
SOURCE_MODE = load_setting("JOBSCRAPER_SOURCES", "linkedin").lower()
OUTPUT_MODE = load_setting("JOBSCRAPER_OUTPUT_MODE", "excel").lower()
EXCEL_OUTPUT_FILE = Path(__file__).with_name("jobs_italy.xlsx")
SPREADSHEET_TITLE = "jobs_italy"

# ─────────────────────────────────────────────
#  KEYWORDS FOR ITALY (mix of English/Italian)
# ─────────────────────────────────────────────

KEYWORDS = [
    # English
    "GIS",
    "GIS Analyst",
    "GIS Developer",
    "Geospatial",
    "Remote Sensing",
    "Earth Observation",
    "Cartography",
    "Geomatics",
    "GeoAI",
    "Geodata",
    "Spatial Data",
    # Italian
    "Sistemi Informativi Geografici",
    "Analista GIS",
    "Sviluppatore GIS",
    "Geomatica",
    "Telerilevamento",
    "Osservazione della Terra",
    "Cartografia",
    "Dati Geospaziali",
    "Geodati",
    "Fotogrammetria",
    "Topografia",
    "Rilievo",
    "Geodesia",
    # Specific roles
    "GIS Specialist",
    "Geospatial Analyst",
    "Mapping",
    "3D Mapping",
    "Urbanistica GIS",
    "Ambiente GIS",
]


# ─────────────────────────────────────────────
#  APIFY API CALL
# ─────────────────────────────────────────────

class ApifyConfigurationError(RuntimeError):
    pass


def apify_headers() -> dict[str, str]:
    return {"Authorization": f"Bearer {APIFY_API_TOKEN}"}


def apify_error_message(response: requests.Response) -> str:
    try:
        data = response.json()
    except ValueError:
        return response.text.strip()[:500] or response.reason
    if isinstance(data, dict):
        error = data.get("error")
        if isinstance(error, dict) and error.get("message"):
            return str(error["message"])
        if data.get("message"):
            return str(data["message"])
    return str(data)[:500]


def build_linkedin_search_url(keyword: str) -> str:
    params = {
        "keywords": keyword,
        "location": LOCATION,
        "geoId": GEO_ID,
        "f_TPR": PUBLISHED_AT,
        "f_E": ",".join(EXPERIENCE_LEVELS),
        "f_JT": ",".join(CONTRACT_TYPES),
        "position": "1",
        "pageNum": "0",
    }
    return f"https://www.linkedin.com/jobs/search/?{urlencode(params)}"


def build_linkedin_actor_input(search_url: str) -> dict:
    return {
        "urls": [search_url],
        "count": MAX_RESULTS_PER_SEARCH,
        "scrapeCompany": SCRAPE_COMPANY_DETAILS,
        "useIncognitoMode": USE_INCOGNITO_MODE,
        "splitByLocation": SPLIT_BY_LOCATION,
    }


def get_searches() -> list[dict]:
    searches = []
    for keyword in KEYWORDS:
        searches.append({
            "source": "linkedin",
            "source_label": "LinkedIn",
            "keyword": keyword,
            "display_label": f"LinkedIn / {keyword}",
            "actor_id": LINKEDIN_ACTOR_ID,
            "payload": build_linkedin_actor_input(build_linkedin_search_url(keyword)),
            "max_items": MAX_RESULTS_PER_SEARCH,
        })
    return searches


def annotate_jobs(jobs: list[dict], source: str, source_label: str) -> list[dict]:
    annotated = []
    for job in jobs:
        job_copy = dict(job)
        job_copy["_source"] = source
        job_copy["_source_label"] = source_label
        annotated.append(job_copy)
    return annotated


def run_actor(actor_id: str, payload: dict, max_items: int) -> list[dict]:
    url = f"https://api.apify.com/v2/acts/{actor_id}/run-sync-get-dataset-items"
    params = {
        "timeout": APIFY_RUN_TIMEOUT_SECONDS,
        "memory": APIFY_RUN_MEMORY_MB,
        "maxItems": max_items,
    }

    response = requests.post(
        url,
        params=params,
        headers=apify_headers(),
        json=payload,
        timeout=APIFY_CLIENT_TIMEOUT_SECONDS,
    )

    if response.status_code in (401, 403):
        raise ApifyConfigurationError(
            f"Apify rejected the request. Check your token and account. "
            f"Apify said: {apify_error_message(response)}"
        )
    response.raise_for_status()
    data = response.json()
    if isinstance(data, list):
        return data
    if isinstance(data, dict) and "items" in data:
        return data["items"]
    return []


def fetch_jobs_for_search(search: dict) -> list[dict]:
    label = search["display_label"]
    try:
        jobs = run_actor(search["actor_id"], search["payload"], search["max_items"])
        return annotate_jobs(jobs, search["source"], search["source_label"])
    except requests.exceptions.Timeout:
        print(f"  ⚠ Timeout for '{label}' — skipping.")
        return []
    except Exception as e:
        print(f"  ⚠ Error for '{label}': {e} — skipping.")
        return []


def run_all_searches(searches: list[dict]) -> list[tuple[str, list]]:
    all_results = []
    print(f"\nRunning up to {SEARCH_CONCURRENCY} search(es) in parallel ...")

    with ThreadPoolExecutor(max_workers=SEARCH_CONCURRENCY) as executor:
        future_to_search = {
            executor.submit(fetch_jobs_for_search, search): search
            for search in searches
        }

        for future in as_completed(future_to_search):
            search = future_to_search[future]
            try:
                jobs = future.result()
                all_results.append((search["keyword"], jobs))
                if jobs:
                    print(f"  ✓ {search['display_label']}: {len(jobs)} job(s)")
                else:
                    print(f"  — {search['display_label']}: 0 results")
            except Exception as e:
                print(f"  ✗ {search['display_label']} failed: {e}")

    return all_results


# ─────────────────────────────────────────────
#  DEDUPLICATION & HELPERS
# ─────────────────────────────────────────────

def make_dedup_key(job: dict) -> str:
    source = str(job.get("_source") or "unknown").lower().strip()
    job_id = job.get("jobId") or job.get("id") or ""
    if job_id:
        return f"{source}|{str(job_id).strip()}"

    title = str(job.get("title", "")).lower().strip()
    company = str(job.get("companyName", "")).lower().strip()
    location = str(job.get("location", "")).lower().strip()
    return f"{source}|{title}|{company}|{location}"


def merge_and_deduplicate(all_results: list[tuple[str, list]]) -> list[dict]:
    seen: dict[str, dict] = {}
    for keyword, jobs in all_results:
        for job in jobs:
            key = make_dedup_key(job)
            if key in seen:
                seen[key]["keywords_matched"].append(keyword)
            else:
                job_copy = dict(job)
                job_copy["keywords_matched"] = [keyword]
                seen[key] = job_copy
    return list(seen.values())


def safe(job: dict, *keys) -> str:
    for k in keys:
        v = job.get(k)
        if v and str(v).strip():
            return str(v).strip()
    return "N/A"


def get_title(job: dict) -> str:
    return safe(job, "title", "jobTitle", "name")


def get_company(job: dict) -> str:
    return safe(job, "companyName", "company")


def get_location(job: dict) -> str:
    return safe(job, "location", "formattedLocation")


def get_job_url(job: dict) -> str:
    url = job.get("jobUrl") or job.get("url") or ""
    if url:
        return url
    job_id = job.get("jobId") or job.get("id") or ""
    if job_id:
        return f"https://www.linkedin.com/jobs/view/{job_id}/"
    return "N/A"


def get_posted(job: dict) -> str:
    posted = job.get("postedAt") or job.get("publishedAt") or ""
    if posted:
        return str(posted)
    return "N/A"


def has_excluded_title(job: dict) -> bool:
    title = get_title(job).casefold()
    return any(term.casefold() in title for term in EXCLUDED_TITLE_TERMS)


def filter_excluded_titles(jobs: list[dict]) -> tuple[list[dict], int]:
    filtered = [job for job in jobs if not has_excluded_title(job)]
    return filtered, len(jobs) - len(filtered)


# ─────────────────────────────────────────────
#  EXCEL EXPORT
# ─────────────────────────────────────────────

HEADER = [
    "Application Status", "Source", "Job Title", "Company", "Location",
    "Job Type", "Posted", "Applicants", "Keywords Matched", "Job URL"
]

COLOR_HEADER_BG = "102C53"
COLOR_HEADER_FG = "FFFFFF"
COLOR_ROW_ODD = "CADCFC"
COLOR_ROW_EVEN = "FFFFFF"

THIN = Side(style="thin", color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def export_to_excel(jobs: list[dict], filename: Path) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = RUN_SHEET_NAME[:31]

    # Create rows
    ws.append(HEADER)
    for job in jobs:
        ws.append([
            "",  # Application Status (manual)
            "LinkedIn",
            get_title(job),
            get_company(job),
            get_location(job),
            safe(job, "employmentType", "jobType"),
            get_posted(job),
            safe(job, "applicantsCount"),
            ", ".join(job.get("keywords_matched", [])),
            get_job_url(job),
        ])

    # Style header
    for cell in ws[1]:
        cell.font = Font(bold=True, color=COLOR_HEADER_FG)
        cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

    # Style data rows
    for row_idx in range(2, ws.max_row + 1):
        bg = COLOR_ROW_ODD if row_idx % 2 == 0 else COLOR_ROW_EVEN
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.border = BORDER
            if col_idx == 10:  # Job URL column
                cell.font = Font(color="0563C1", underline="single")

    # Set column widths
    widths = [18, 12, 40, 30, 25, 18, 20, 12, 35, 50]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A2"
    wb.save(filename)
    return str(filename)


# ─────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────

def main():
    run_started = time.perf_counter()

    # Check token
    if not APIFY_API_TOKEN or APIFY_API_TOKEN == TOKEN_PLACEHOLDER:
        print(f"❌ Please set {TOKEN_ENV_VAR} in .env file")
        print(f"   Create .env with: {TOKEN_ENV_VAR}=your_token_here")
        return

    searches = get_searches()
    print("=" * 60)
    print(f"  LinkedIn Job Scraper — ITALY")
    print(f"  Run started: {RUN_STARTED_AT.strftime('%Y-%m-%d %H:%M %Z')}")
    print(f"  Location: {LOCATION}")
    print(f"  Keywords: {len(KEYWORDS)}")
    print(f"  Max results per keyword: {MAX_RESULTS_PER_SEARCH}")
    print(f"  Search concurrency: {SEARCH_CONCURRENCY}")
    print("=" * 60)

    # Run searches
    all_results = run_all_searches(searches)

    # Deduplicate
    print("\n" + "─" * 60)
    print("Deduplicating results ...")
    unique_jobs = merge_and_deduplicate(all_results)
    print(f"  → {len(unique_jobs)} unique job(s)")

    # Filter excluded titles
    unique_jobs, excluded = filter_excluded_titles(unique_jobs)
    if excluded:
        print(f"  → Removed {excluded} job(s) with excluded terms")

    # Export
    output_path = export_to_excel(unique_jobs, EXCEL_OUTPUT_FILE)

    # Summary
    print("\n" + "=" * 60)
    print(f"  ✓ Completed! Found {len(unique_jobs)} unique jobs.")
    print(f"  Excel file: {output_path}")
    print(f"  Runtime: {time.perf_counter() - run_started:.1f} seconds")
    print("=" * 60)


if __name__ == "__main__":
    main()